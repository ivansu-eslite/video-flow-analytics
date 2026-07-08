"""Zone 人流 Excel 報表：離線下游步驟（CLI 的 `report` 子命令）。

讀 `outputs/{bucket}/{date}/zone_counts.parquet`，彙總成跨日累加更新的
`outputs/{bucket}/report.xlsx`。實際的期間彙總／尖峰計算在 report/stats.py；
這裡負責讀檔、驗證、orchestration 與 Excel 讀寫。
"""

import datetime
import logging
from pathlib import Path
from typing import Literal

import openpyxl
import polars as pl
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from video_flow_analytics.core.config import settings
from video_flow_analytics.core.registry import (
    CameraRegistry,
    _find_duplicates,
    load_registry_from_path,
)
from video_flow_analytics.report.stats import peak_per_day, rollup_by_period, to_taipei

logger = logging.getLogger(__name__)

OUTPUT_ROOT = Path("outputs")


def _validate_unique_zone_names(registry: CameraRegistry) -> None:
    """報表以 zone 名稱（不含 camera_id）分組，因此要求整份 registry 的 zone
    名稱全域唯一；此驗證只在產報表時檢查，不影響 analyze_daily / zone_mapping。
    """
    names = [zone.name for cam in registry.cameras for zone in cam.parsed_zones()]
    dupes = sorted(_find_duplicates(names))
    if dupes:
        raise ValueError(
            "camera_registry.yaml 中有跨攝影機重複的 zone 名稱，報表需要 zone "
            f"名稱全域唯一（不只同一攝影機內唯一）: {dupes}"
        )


def _build_report_frames(
    date: datetime.date,
    bucket_dir: str,
    period_minutes: int,
    metric: str,
    bucket_minutes: int,
    output_root: Path = OUTPUT_ROOT,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    if period_minutes % bucket_minutes != 0:
        raise ValueError(
            f"report.period_minutes（{period_minutes}）必須是 "
            f"zone.bucket_minutes（{bucket_minutes}）的倍數。"
        )

    bucket_path = Path(bucket_dir)
    output_dir = output_root / bucket_path.name / date.isoformat()
    counts_path = output_dir / "zone_counts.parquet"
    if not counts_path.exists():
        raise FileNotFoundError(
            f"找不到 zone 人流統計 {counts_path}，"
            "請先執行 map_zones_daily 產生當日 parquet。"
        )

    # 驗證用產生 zone_counts.parquet 當時的 registry 快照，而非「當下」的 registry，
    # 否則兩者間若改過 zone 名稱，可能讓不同攝影機的人流被靜默合併
    registry = load_registry_from_path(output_dir / "camera_registry_used.yaml")
    _validate_unique_zone_names(registry)

    df = to_taipei(pl.read_parquet(counts_path))
    hourly_df = rollup_by_period(df, period_minutes, metric)
    peak_df = peak_per_day(hourly_df)
    return hourly_df, peak_df


SHEET_HOURLY = "每小時人流"
SHEET_PEAK = "每日尖峰"
SHEET_EVENTS = "活動事件"

_HOURLY_HEADERS = ["日期", "星期", "小時", "區域", "人流量"]
_PEAK_HEADERS = ["日期", "星期", "區域", "尖峰時段", "尖峰人流", "每日提醒"]
_EVENTS_HEADERS = [
    "日期",
    "星期",
    "開始時間",
    "結束時間",
    "區域",
    "活動名稱",
    "活動類型",
]

# _sort_rows 用欄位名稱指定排序鍵，避免欄位順序調整時忘記同步改數字索引。
_HOURLY_SORT_COLUMNS = ("日期", "小時", "區域")
_PEAK_SORT_COLUMNS = ("日期", "區域")


def _sort_key_columns(headers: list[str], columns: tuple[str, ...]) -> tuple[int, ...]:
    return tuple(headers.index(column) for column in columns)


def _init_sheet(wb: Workbook, name: str, headers: list[str]) -> Worksheet:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 14
    return ws


def _existing_dates(ws: Worksheet) -> set[str]:
    return {row[0].value for row in ws.iter_rows(min_row=2) if row[0].value is not None}


def _remove_rows_for_dates(ws: Worksheet, dates: set[str]) -> None:
    rows_to_delete = [
        row[0].row for row in ws.iter_rows(min_row=2) if row[0].value in dates
    ]
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)


def _append_rows(ws: Worksheet, df: pl.DataFrame) -> None:
    for row in df.iter_rows():
        ws.append(row)


def _sort_rows(ws: Worksheet, key_columns: tuple[int, ...]) -> None:
    if ws.max_row < 2:
        return
    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    rows.sort(key=lambda r: tuple(r[i] for i in key_columns))
    ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append(row)


def _write_report(
    path: Path,
    hourly_new: pl.DataFrame,
    peak_new: pl.DataFrame,
    on_duplicate_date: Literal["overwrite", "append", "error"],
) -> None:
    new_dates = set(hourly_new["date"].to_list())

    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        _init_sheet(wb, SHEET_HOURLY, _HOURLY_HEADERS)
        _init_sheet(wb, SHEET_PEAK, _PEAK_HEADERS)
        _init_sheet(wb, SHEET_EVENTS, _EVENTS_HEADERS)

    hourly_ws = wb[SHEET_HOURLY]
    peak_ws = wb[SHEET_PEAK]

    if on_duplicate_date == "error":
        conflict = new_dates & (_existing_dates(hourly_ws) | _existing_dates(peak_ws))
        if conflict:
            raise ValueError(
                f"報表中已存在這些日期的資料，未寫入任何內容：{sorted(conflict)}"
                "（可改用 on_duplicate_date='overwrite' 或 'append'）"
            )

    if on_duplicate_date == "overwrite":
        _remove_rows_for_dates(hourly_ws, new_dates)
        _remove_rows_for_dates(peak_ws, new_dates)

    _append_rows(hourly_ws, hourly_new)
    _append_rows(peak_ws, peak_new)

    if on_duplicate_date == "overwrite":
        _sort_rows(
            hourly_ws,
            key_columns=_sort_key_columns(_HOURLY_HEADERS, _HOURLY_SORT_COLUMNS),
        )
        _sort_rows(
            peak_ws, key_columns=_sort_key_columns(_PEAK_HEADERS, _PEAK_SORT_COLUMNS)
        )

    tmp_path = path.with_name(path.name + ".tmp")
    wb.save(tmp_path)
    tmp_path.replace(path)


def export_report_daily(
    date: datetime.date,
    bucket_dir: str,
    period_minutes: int,
    metric: Literal["entries", "unique_visitors"],
    on_duplicate_date: Literal["overwrite", "append", "error"],
    bucket_minutes: int,
    output_root: Path = OUTPUT_ROOT,
) -> Path:
    """執行單日 zone 人流報表彙總，寫入跨日累加更新的 `report.xlsx`。

    純 CPU 運算，不需重跑偵測或 zone mapping；讀取的 registry 是產生
    `zone_counts.parquet` 當時的 `camera_registry_used.yaml` 快照（而非
    當下的 `camera_registry.yaml`），以避免同名 zone 被靜默合併。

    Args:
        date: 要彙總的日期，需已有對應的 `zone_counts.parquet`。
        bucket_dir: 本機模擬 GCS bucket 的根目錄。
        period_minutes: 報表人流彙總的時段粒度（分鐘），需為 `bucket_minutes`
            的倍數。
        metric: 「人流量」「尖峰人流」使用的統計量。
        on_duplicate_date: 同一天資料已存在時的處理方式。
        bucket_minutes: `zone_counts.parquet` 的時段粒度（分鐘）。
        output_root: 輸出根目錄。

    Returns:
        `report.xlsx` 的路徑。

    Raises:
        ValueError: `period_minutes` 不是 `bucket_minutes` 的倍數、
            `camera_registry_used.yaml` 中有跨攝影機重複的 zone 名稱，或
            `on_duplicate_date="error"` 時發現日期已存在。
        FileNotFoundError: 當日 `zone_counts.parquet` 不存在。
    """
    hourly_df, peak_df = _build_report_frames(
        date, bucket_dir, period_minutes, metric, bucket_minutes, output_root
    )

    bucket_name = Path(bucket_dir).name
    report_path = output_root / bucket_name / "report.xlsx"
    _write_report(report_path, hourly_df, peak_df, on_duplicate_date)

    logger.info(
        "Zone 人流報表已寫入 %s（本次日期：%s，%d 列每小時人流、%d 列每日尖峰）。",
        report_path,
        sorted(hourly_df["date"].unique().to_list()),
        hourly_df.height,
        peak_df.height,
    )
    return report_path


def run_report() -> None:
    """`report` 子命令的進入點：從 `config.toml` 取參數後呼叫 `export_report_daily`。

    Raises:
        ValueError: `config.toml` 的 `[input].date` 未設定。
    """
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )
    if settings.input.date is None:
        raise ValueError("config.toml 的 [input].date 未設定，請指定要分析的日期。")

    export_report_daily(
        date=settings.input.date,
        bucket_dir=settings.input.bucket_dir,
        period_minutes=settings.report.period_minutes,
        metric=settings.report.metric,
        on_duplicate_date=settings.report.on_duplicate_date,
        bucket_minutes=settings.zone.bucket_minutes,
    )


if __name__ == "__main__":
    run_report()
