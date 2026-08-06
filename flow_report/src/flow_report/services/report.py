"""人流 Excel 報表：核心匯出邏輯（讀檔、驗證、orchestration 與 Excel 讀寫）。

讀 `outputs/{bucket}/{date}/` 下的 `zone_counts.parquet`（區域佔用）與
`line_counts.parquet`（計數線進出），彙總成跨日累加更新的
`outputs/{bucket}/report.xlsx`。實際的期間彙總／尖峰計算在 `services/stats.py`。

**哪些輸入是必要的，由 `bucket_dir/camera_registry.yaml` 的定義決定**，不是「檔案在
不在」——後者無法區分「定義了計數線卻忘了跑 `line_counting`」與「這個 bucket 本來
就沒有計數線」，前者會靜默少三頁。見 ADR-005、ADR-007。
"""

import datetime
from pathlib import Path
from typing import Literal, NamedTuple

import openpyxl
import polars as pl
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from vfa_observability import StructuredLogger
from vfa_registry import (
    CameraEntry,
    load_registry,
    parse_and_validate_lines,
    parse_and_validate_zones,
    registry_path,
)

from flow_report.config.constants import (
    COLUMN_WIDTH,
    EVENTS_HEADERS,
    LINE_COUNTS_FILENAME,
    LINE_HOURLY_HEADERS,
    LINE_HOURLY_SORT_COLUMNS,
    LINE_PEAK_HEADERS,
    LINE_PEAK_SORT_COLUMNS,
    OUTPUT_ROOT,
    REPORT_FILENAME,
    SHEET_EVENTS,
    SHEET_LINE_HOURLY,
    SHEET_LINE_PEAK_IN,
    SHEET_LINE_PEAK_OUT,
    SHEET_ZONE_HOURLY,
    SHEET_ZONE_PEAK,
    TMP_SUFFIX,
    ZONE_COUNTS_FILENAME,
    ZONE_HOURLY_HEADERS,
    ZONE_HOURLY_SORT_COLUMNS,
    ZONE_PEAK_HEADERS,
    ZONE_PEAK_SORT_COLUMNS,
)
from flow_report.services.stats import (
    peak_lines_per_day,
    peak_per_day,
    rollup_by_period,
    rollup_lines_by_period,
    to_taipei,
)

logger = StructuredLogger(component="report_builder")


class ReportFrames(NamedTuple):
    """本次要寫入報表的各分頁資料；`None` 代表該類統計在這個 bucket 沒有定義。

    `None` 與「空的 DataFrame」語義不同：前者是 registry 裡根本沒有區域／計數線定義，
    本次不產生該類資料（分頁仍會建立、`overwrite` 時仍會清掉該日舊列）；後者是
    有定義但當日沒有事件，是上游的正常產物。
    """

    zone_hourly: pl.DataFrame | None
    zone_peak: pl.DataFrame | None
    line_hourly: pl.DataFrame | None
    line_peak_in: pl.DataFrame | None
    line_peak_out: pl.DataFrame | None


class _DataSheet(NamedTuple):
    """一個由本階段寫入的分頁：分頁名、表頭、排序欄與對應的 `ReportFrames` 欄位。"""

    name: str
    headers: list[str]
    sort_columns: tuple[str, ...]
    field: str


# 由本階段寫入的分頁。`活動事件` 不在此列：它的寫入者是其他來源，本階段只建表頭，
# 既不寫入、也不因 overwrite 而清除該日的列（清了會刪掉別人的資料）。
_DATA_SHEETS = (
    _DataSheet(
        SHEET_ZONE_HOURLY, ZONE_HOURLY_HEADERS, ZONE_HOURLY_SORT_COLUMNS, "zone_hourly"
    ),
    _DataSheet(SHEET_ZONE_PEAK, ZONE_PEAK_HEADERS, ZONE_PEAK_SORT_COLUMNS, "zone_peak"),
    _DataSheet(
        SHEET_LINE_HOURLY, LINE_HOURLY_HEADERS, LINE_HOURLY_SORT_COLUMNS, "line_hourly"
    ),
    _DataSheet(
        SHEET_LINE_PEAK_IN, LINE_PEAK_HEADERS, LINE_PEAK_SORT_COLUMNS, "line_peak_in"
    ),
    _DataSheet(
        SHEET_LINE_PEAK_OUT, LINE_PEAK_HEADERS, LINE_PEAK_SORT_COLUMNS, "line_peak_out"
    ),
)

# 報表的完整分頁清單（含只建表頭的 `活動事件`）；順序即分頁在檔案中的順序。
_SHEET_LAYOUT = tuple(
    [(sheet.name, sheet.headers) for sheet in _DATA_SHEETS]
    + [(SHEET_EVENTS, EVENTS_HEADERS)]
)


def _reject_unknown_pairs(
    df: pl.DataFrame,
    columns: tuple[str, str],
    valid_pairs: set[tuple[str, str]],
    source_path: Path,
) -> None:
    """parquet 出現不在 registry 定義內的組合時 fail-loud，而非靜默讀入未經驗證的資料。"""
    actual_pairs = set(df.select(list(columns)).unique(maintain_order=True).iter_rows())
    unknown_pairs = actual_pairs - valid_pairs
    if unknown_pairs:
        raise ValueError(
            f"{source_path} 出現不在 camera_registry.yaml 定義內的 "
            f"({columns[0]}, {columns[1]}) 組合: {sorted(unknown_pairs)}"
        )


def _reject_orphan_counts(counts_path: Path, registry_file: Path, kind: str) -> None:
    """registry 已無該側定義、當日 parquet 卻有資料時 fail-loud。

    這是「registry 定義了卻沒跑上游」的反向情況：整側定義被清空（或該側所有攝影機
    的 `participates_in_zone_mapping` 被關掉），但當日 parquet 還帶著用舊定義算出來
    的資料。靜默跳過的話，該類統計會整批從報表消失，且 `on_duplicate_date`
    ＝`overwrite` 時連既有的舊列一併清掉——正是這條 repo 要擋的那種無聲資料損失。

    `_reject_unknown_pairs` 只擋得住「部分定義被移除」（其餘定義還在，該側仍會進到
    這裡的下游）；整側清空時那條路根本走不到，才需要這道檢查。

    0 列的 parquet 不算：那是上游對「這個 bucket 沒有該側定義」的正常產物
    （執行了、沒有東西可算），不是錯位。
    """
    if not counts_path.exists() or pl.read_parquet(counts_path).height == 0:
        return
    raise ValueError(
        f"{registry_file} 中已沒有任何攝影機定義{kind}，但 {counts_path} 仍有資料。"
        f"若是誤刪定義，請把{kind}定義補回 registry；若確定不再統計{kind}，"
        "請一併移除該日的 parquet（本階段不會自行刪除上游產物）。"
    )


def _zone_frames(
    output_dir: Path,
    zone_entries: dict[str, CameraEntry],
    period_minutes: int,
    metric: str,
    registry_file: Path,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    counts_path = output_dir / ZONE_COUNTS_FILENAME
    if not counts_path.exists():
        raise FileNotFoundError(
            f"{registry_file} 中有攝影機定義了區域，但找不到區域事件統計 "
            f"{counts_path}，請先執行 map_zones_daily 產生當日 parquet。"
        )

    # parse_and_validate_zones 順便驗證跨攝影機 zone 名稱唯一性
    zone_cameras = parse_and_validate_zones(zone_entries)
    df = to_taipei(pl.read_parquet(counts_path))
    valid_pairs = {
        (camera_id, zone.name)
        for camera_id, zones in zone_cameras.items()
        for zone in zones
    }
    _reject_unknown_pairs(df, ("camera_id", "zone"), valid_pairs, counts_path)

    hourly_df = rollup_by_period(df, period_minutes, metric)
    return hourly_df, peak_per_day(hourly_df)


def _line_frames(
    output_dir: Path,
    line_entries: dict[str, CameraEntry],
    period_minutes: int,
    registry_file: Path,
) -> tuple[pl.DataFrame, pl.DataFrame, pl.DataFrame]:
    counts_path = output_dir / LINE_COUNTS_FILENAME
    if not counts_path.exists():
        raise FileNotFoundError(
            f"{registry_file} 中有攝影機定義了計數線，但找不到計數線進出統計 "
            f"{counts_path}，請先執行 count_lines_daily 產生當日 parquet。"
        )

    # parse_and_validate_lines 順便驗證跨攝影機 line 名稱唯一性
    line_cameras = parse_and_validate_lines(line_entries)
    df = to_taipei(pl.read_parquet(counts_path))
    valid_pairs = {
        (camera_id, line.name)
        for camera_id, lines in line_cameras.items()
        for line in lines
    }
    _reject_unknown_pairs(df, ("camera_id", "line"), valid_pairs, counts_path)

    hourly_df = rollup_lines_by_period(df, period_minutes)
    return (
        hourly_df,
        peak_lines_per_day(hourly_df, "in_count"),
        peak_lines_per_day(hourly_df, "out_count"),
    )


def _build_report_frames(
    date: datetime.date,
    bucket_dir: str,
    period_minutes: int,
    metric: str,
    bucket_minutes: int,
    output_root: Path = OUTPUT_ROOT,
) -> ReportFrames:
    if period_minutes % bucket_minutes != 0:
        raise ValueError(
            f"report.period_minutes（{period_minutes}）必須是 "
            f"input.bucket_minutes（{bucket_minutes}）的倍數。"
        )

    bucket_path = Path(bucket_dir)
    output_dir = output_root / bucket_path.name / date.isoformat()
    registry_file = registry_path(bucket_path)
    registry = load_registry(bucket_path)

    # zone 名稱唯一性的驗證範圍維持既有的 participates_in_zone_mapping 篩選（不加
    # 「zones 非空」條件）；「該不該有 zone_counts.parquet」是另一個判斷，才看 zones。
    zone_entries = {
        entry.stream_dirname: entry
        for entry in registry.cameras
        if entry.participates_in_zone_mapping
    }
    # 計數線沒有對應的參與旗標，`lines` 非空即代表參與（同 line_counting）。
    line_entries = {
        entry.stream_dirname: entry for entry in registry.cameras if entry.lines
    }
    has_zone_defs = any(entry.zones for entry in zone_entries.values())
    has_line_defs = bool(line_entries)
    if not has_zone_defs and not has_line_defs:
        raise ValueError(
            f"{registry_file} 中沒有任何攝影機定義區域或計數線，沒有可彙總的統計。"
        )

    zone_hourly = zone_peak = None
    if has_zone_defs:
        zone_hourly, zone_peak = _zone_frames(
            output_dir, zone_entries, period_minutes, metric, registry_file
        )
    else:
        _reject_orphan_counts(
            output_dir / ZONE_COUNTS_FILENAME, registry_file, "區域"
        )

    line_hourly = line_peak_in = line_peak_out = None
    if has_line_defs:
        line_hourly, line_peak_in, line_peak_out = _line_frames(
            output_dir, line_entries, period_minutes, registry_file
        )
    else:
        _reject_orphan_counts(
            output_dir / LINE_COUNTS_FILENAME, registry_file, "計數線"
        )

    return ReportFrames(
        zone_hourly=zone_hourly,
        zone_peak=zone_peak,
        line_hourly=line_hourly,
        line_peak_in=line_peak_in,
        line_peak_out=line_peak_out,
    )


def _sort_key_columns(headers: list[str], columns: tuple[str, ...]) -> tuple[int, ...]:
    return tuple(headers.index(column) for column in columns)


def _init_sheet(wb: Workbook, name: str, headers: list[str]) -> Worksheet:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = COLUMN_WIDTH
    return ws


def _cell_date_str(value: object) -> str | None:
    """把日期欄的儲格值正規化成 `YYYY-MM-DD` 字串。

    本階段以字串寫入日期，但 Excel／BI 工具存檔時可能把該欄轉成日期型別的儲格，
    讀回時即為 `datetime.date`／`datetime.datetime`。日期欄同時是比對與排序的鍵，
    型別混雜會讓比對永遠不成立、排序直接拋 `TypeError`，故一律正規化後再使用。
    """
    if value is None:
        return None
    if isinstance(value, datetime.date):  # datetime.datetime 亦為其子類
        return value.strftime("%Y-%m-%d")
    return str(value)


def _existing_dates(ws: Worksheet) -> set[str]:
    dates = (_cell_date_str(row[0].value) for row in ws.iter_rows(min_row=2))
    return {date for date in dates if date is not None}


def _remove_rows_for_dates(ws: Worksheet, dates: set[str]) -> None:
    rows_to_delete = [
        row[0].row
        for row in ws.iter_rows(min_row=2)
        if _cell_date_str(row[0].value) in dates
    ]
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)


def _append_rows(ws: Worksheet, df: pl.DataFrame) -> None:
    for row in df.iter_rows():
        ws.append(row)


def _sort_key(value: object) -> object:
    """排序鍵正規化：`None` 與日期型別的儲格轉成字串，其餘型別原樣保留。

    key_columns 可能包含非日期欄（如區域名稱），只正規化日期型別可避免混入
    `datetime.date`／`str` 時排序互相比較拋 `TypeError`，同時不影響其他欄位
    的原生型別比較。`None` 同理：既有 report.xlsx 的分頁可能含全空列，拿 `None`
    與 `str` 比較一樣會 `TypeError`，這裡回空字串讓它穩定排在最前面（本階段的
    排序欄都是字串欄）。
    """
    if value is None:
        return ""
    if isinstance(value, datetime.date):  # datetime.datetime 亦為其子類
        return _cell_date_str(value)
    return value


def _sort_rows(ws: Worksheet, key_columns: tuple[int, ...]) -> None:
    if ws.max_row < 2:
        return
    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    rows.sort(key=lambda r: tuple(_sort_key(r[i]) for i in key_columns))
    ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append(row)


def _frame_dates(frames: ReportFrames, date: datetime.date) -> set[str]:
    """本次涉及的日期：本次彙總的 `date`，加上各分頁資料實際帶到的日期。

    取聯集而非各分頁各自為政，是為了讓 `overwrite` 維持「該日的完整重寫」這個
    單一語義——否則 registry 移除 `lines` 後重跑該日，出入口三頁會留著舊列、
    區域兩頁換成新列，同一天在不同分頁混雜新舊資料。

    `date` 一律計入，資料本身可能一列都沒有（上游重跑後該日事件清空是 0 列，
    不是缺資料），只看資料內容的話這種情況會清不到任何列，該日的舊資料留在
    報表裡，同樣不符「完整重寫」。
    """
    dates: set[str] = {date.isoformat()}
    for sheet in _DATA_SHEETS:
        df = getattr(frames, sheet.field)
        if df is not None:
            dates |= set(df["date"].to_list())
    return dates


def _write_report(
    path: Path,
    frames: ReportFrames,
    date: datetime.date,
    on_duplicate_date: Literal["overwrite", "append", "error"],
) -> None:
    new_dates = _frame_dates(frames, date)

    if path.exists():
        wb = openpyxl.load_workbook(path)
        default_sheet = None
    else:
        wb = Workbook()
        default_sheet = wb.active

    try:
        # 缺哪個分頁就補建哪個：既有檔可能是本次改名前的舊格式，也可能還沒有出入口
        # 三頁。一律先查 `sheetnames`——`create_sheet` 對同名分頁不報錯，會靜默改名成
        # `活動事件1`，直接呼叫 `_init_sheet` 會在既有檔上生出重複分頁。
        for name, headers in _SHEET_LAYOUT:
            if name not in wb.sheetnames:
                _init_sheet(wb, name, headers)
        if default_sheet is not None:
            wb.remove(default_sheet)

        if on_duplicate_date == "error":
            existing_dates: set[str] = set()
            for sheet in _DATA_SHEETS:
                existing_dates |= _existing_dates(wb[sheet.name])
            conflict = new_dates & existing_dates
            if conflict:
                # 在 wb.save 之前拋出，本次不寫入任何分頁（跨分頁的原子性）
                raise ValueError(
                    f"報表中已存在這些日期的資料，未寫入任何內容：{sorted(conflict)}"
                    "（可改用 on_duplicate_date='overwrite' 或 'append'）"
                )

        for sheet in _DATA_SHEETS:
            ws = wb[sheet.name]
            df = getattr(frames, sheet.field)
            if on_duplicate_date == "overwrite":
                # `None` 的分頁只清不寫，見 _frame_dates 的說明
                _remove_rows_for_dates(ws, new_dates)
            if df is not None:
                _append_rows(ws, df)
            if on_duplicate_date == "overwrite":
                _sort_rows(
                    ws,
                    key_columns=_sort_key_columns(sheet.headers, sheet.sort_columns),
                )

        tmp_path = path.with_name(path.name + TMP_SUFFIX)
        wb.save(tmp_path)
        tmp_path.replace(path)
    finally:
        wb.close()


def export_report_daily(
    date: datetime.date,
    bucket_dir: str,
    period_minutes: int,
    metric: Literal["entries", "unique_visitors"],
    on_duplicate_date: Literal["overwrite", "append", "error"],
    bucket_minutes: int,
    output_root: Path = OUTPUT_ROOT,
) -> Path:
    """執行單日人流報表彙總，寫入跨日累加更新的 `report.xlsx`。

    純 CPU 運算，不需重跑偵測、zone mapping 或計數線統計；讀取的 registry 是
    `bucket_dir` 底下當下的 `camera_registry.yaml`（見 ADR-007）。

    **registry 同時決定哪些輸入是必要的**：有攝影機定義了區域就必須有
    `zone_counts.parquet`、有攝影機定義了計數線就必須有 `line_counts.parquet`，
    缺檔即報錯；registry 裡沒有定義的那一側則整批跳過（分頁仍建立、不寫入資料），
    不算錯誤。兩者都沒有定義則沒有東西可彙總，直接報錯。詳見 ADR-005。

    Args:
        date: 要彙總的日期，需已有 registry 所要求的當日 parquet。
        bucket_dir: 本機模擬 GCS bucket 的根目錄。
        period_minutes: 報表人流彙總的時段粒度（分鐘），需為 `bucket_minutes`
            的倍數。
        metric: 「人流量」「尖峰人流」使用的統計量；只作用於區域統計，計數線
            固定用 `in_count`／`out_count`。
        on_duplicate_date: 同一天資料已存在時的處理方式。
        bucket_minutes: 上游 parquet 的時段粒度（分鐘）。
        output_root: 輸出根目錄。

    Returns:
        `report.xlsx` 的路徑。

    Raises:
        ValueError: `period_minutes` 不是 `bucket_minutes` 的倍數、
            `camera_registry.yaml` 中有跨攝影機重複的 zone／line 名稱、
            registry 中沒有任何區域或計數線定義、上游 parquet 出現不在該 registry
            內的 (camera_id, zone)／(camera_id, line) 組合、registry 已無某一側的
            定義但當日對應 parquet 仍有資料，或 `on_duplicate_date="error"` 時
            發現日期已存在。
        FileNotFoundError: registry 要求的當日 parquet 不存在，或 `bucket_dir`
            底下找不到 `camera_registry.yaml`。
    """
    frames = _build_report_frames(
        date, bucket_dir, period_minutes, metric, bucket_minutes, output_root
    )

    bucket_name = Path(bucket_dir).name
    report_path = output_root / bucket_name / REPORT_FILENAME
    _write_report(report_path, frames, date, on_duplicate_date)

    logger.info(
        "人流報表已寫入",
        path=str(report_path),
        dates=sorted(_frame_dates(frames, date)),
        rows_by_sheet={
            sheet.name: (
                None
                if getattr(frames, sheet.field) is None
                else getattr(frames, sheet.field).height
            )
            for sheet in _DATA_SHEETS
        },
    )
    return report_path
