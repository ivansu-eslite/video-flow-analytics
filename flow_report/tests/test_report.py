import datetime
from pathlib import Path
from unittest import mock
from zoneinfo import ZoneInfo

import openpyxl
import polars as pl
import pytest
import yaml

from flow_report.config.constants import (
    EVENTS_HEADERS,
    LINE_HOURLY_HEADERS,
    LINE_PEAK_HEADERS,
    SHEET_EVENTS,
    SHEET_LINE_HOURLY,
    SHEET_LINE_PEAK_IN,
    SHEET_LINE_PEAK_OUT,
    SHEET_ZONE_HOURLY,
    SHEET_ZONE_PEAK,
    ZONE_HOURLY_HEADERS,
    ZONE_PEAK_HEADERS,
)
from flow_report.services.report import (
    ReportFrames,
    _build_report_frames,
    _write_report,
)

_ALL_SHEETS = (
    SHEET_ZONE_HOURLY,
    SHEET_ZONE_PEAK,
    SHEET_LINE_HOURLY,
    SHEET_LINE_PEAK_IN,
    SHEET_LINE_PEAK_OUT,
    SHEET_EVENTS,
)


def _write_registry(
    path: Path,
    zones_by_camera: dict[str, list[str]] | None = None,
    lines_by_camera: dict[str, list[str]] | None = None,
) -> None:
    """寫出 registry；zones／lines 皆以 camera_id -> 名稱清單 指定。"""
    zones_by_camera = zones_by_camera or {}
    lines_by_camera = lines_by_camera or {}
    camera_ids = sorted({*zones_by_camera, *lines_by_camera})
    data = {
        "bucket_name": "bucket_test",
        "storage": {},
        "cameras": [
            {
                "camera_id": cam_id,
                "location": "loc",
                "ip": "127.0.0.1",
                "zones": [
                    {"name": name, "polygon": [[0, 0], [1, 0], [1, 1]]}
                    for name in zones_by_camera.get(cam_id, [])
                ],
                "lines": [
                    {
                        "name": name,
                        "points": [[0, 0], [10, 0]],
                        "inside_point": [5, 5],
                        "line_group": "四樓書店",
                    }
                    for name in lines_by_camera.get(cam_id, [])
                ],
            }
            for cam_id in camera_ids
        ],
    }
    path.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8"
    )


def _write_zone_counts(
    path: Path, camera_id: str = "loc_cam001", zone: str = "entrance"
) -> None:
    df = pl.DataFrame(
        {
            "camera_id": [camera_id],
            "zone": [zone],
            "time_bucket": [
                datetime.datetime(2026, 5, 1, 11, 0, tzinfo=ZoneInfo("Asia/Taipei"))
            ],
            "unique_visitors": [1],
            "entries": [1],
        },
        schema={
            "camera_id": pl.Utf8,
            "zone": pl.Utf8,
            "time_bucket": pl.Datetime("us", "Asia/Taipei"),
            "unique_visitors": pl.Int64,
            "entries": pl.Int64,
        },
    )
    df.write_parquet(path)


def _write_line_counts(
    path: Path,
    camera_id: str = "loc_cam001",
    line: str = "main_gate",
    empty: bool = False,
) -> None:
    rows = (
        {
            "line_group": [],
            "camera_id": [],
            "line": [],
            "time_bucket": [],
            "in_count": [],
            "out_count": [],
        }
        if empty
        else {
            "line_group": ["四樓書店"],
            "camera_id": [camera_id],
            "line": [line],
            "time_bucket": [
                datetime.datetime(2026, 5, 1, 11, 0, tzinfo=ZoneInfo("Asia/Taipei"))
            ],
            "in_count": [7],
            "out_count": [4],
        }
    )
    pl.DataFrame(
        rows,
        schema={
            "line_group": pl.Utf8,
            "camera_id": pl.Utf8,
            "line": pl.Utf8,
            "time_bucket": pl.Datetime("us", "Asia/Taipei"),
            "in_count": pl.Int64,
            "out_count": pl.Int64,
        },
    ).write_parquet(path)


def _prepare_bucket(tmp_path: Path) -> tuple[Path, Path]:
    """建出 bucket 與當日輸出目錄，回傳 (bucket_dir, output_dir)。"""
    bucket_dir = tmp_path / "bucket_test"
    bucket_dir.mkdir()
    output_dir = tmp_path / "outputs" / "bucket_test" / "2026-05-01"
    output_dir.mkdir(parents=True)
    return bucket_dir, output_dir


def _build(tmp_path: Path, bucket_dir: Path, **kwargs) -> ReportFrames:
    params = {
        "date": datetime.date(2026, 5, 1),
        "bucket_dir": str(bucket_dir),
        "period_minutes": 60,
        "metric": "entries",
        "bucket_minutes": 15,
        "output_root": tmp_path / "outputs",
    }
    params.update(kwargs)
    return _build_report_frames(**params)


def test_build_report_frames_uses_live_registry_not_leftover_snapshot(tmp_path):
    """registry 來源是 `bucket_dir` 當下的 camera_registry.yaml；舊版留在輸出目錄
    下的 camera_registry_used.yaml 不再被讀取（見 ADR-007）。

    刻意讓兩份內容不同：當下的 registry 只有 cam001，殘留快照多了一台 cam002。
    若還在讀快照，(loc_cam002, lobby) 這組會被視為合法而不報錯。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"]},
    )
    # 舊版遺留的快照：內容與當下的 registry 不同，不該影響結果
    _write_registry(
        output_dir / "camera_registry_used.yaml",
        {"cam001": ["entrance"], "cam002": ["lobby"]},
    )
    _write_zone_counts(
        output_dir / "zone_counts.parquet", camera_id="loc_cam002", zone="lobby"
    )

    with pytest.raises(ValueError, match="不在.*定義"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_requires_live_registry(tmp_path):
    """`bucket_dir` 下沒有 camera_registry.yaml 時 fail-loud。

    上游 parquet 都在也不能退回「看檔案在不在」——那正是 ADR-005 要避免的靜默少頁。
    """
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_zone_counts(output_dir / "zone_counts.parquet")

    # 路徑一併釘住：舊實作缺快照時拋的是同一句話、只有檔名不同，不比對檔名的話
    # 這支測試在改回讀快照後仍會是綠的
    with pytest.raises(FileNotFoundError, match=r"找不到設備登錄檔.*camera_registry\.yaml$"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_rejects_orphan_zone_counts(tmp_path):
    """registry 已無任何區域定義、當日 zone_counts.parquet 卻有資料時要擋下。

    `_reject_unknown_pairs` 只擋得住「部分 zone 被移除」；整側清空時那條路走不到，
    靜默跳過會讓區域統計整批從報表消失，`overwrite` 重跑還會清掉既有的舊列。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        lines_by_camera={"cam001": ["main_gate"]},
    )
    _write_line_counts(output_dir / "line_counts.parquet")
    _write_zone_counts(output_dir / "zone_counts.parquet")

    with pytest.raises(ValueError, match="已沒有任何攝影機定義區域"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_accepts_orphan_but_empty_counts(tmp_path):
    """0 列的 parquet 不算錯位：那是上游對「這個 bucket 沒有該側定義」的正常產物
    （執行了、沒有東西可算），不該把它當成 registry 被誤改。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"]},
    )
    _write_zone_counts(output_dir / "zone_counts.parquet")
    _write_line_counts(output_dir / "line_counts.parquet", empty=True)

    frames = _build(tmp_path, bucket_dir)
    assert frames.zone_hourly.height == 1
    assert frames.line_hourly is None


def test_build_report_frames_rejects_live_registry_duplicates(tmp_path):
    """當下的 camera_registry.yaml 有跨攝影機重複的 zone 名稱時要擋下。

    改讀當下 registry 後，這正是要 fail-loud 的情況：報表以 zone 名稱（不含
    camera_id）分組，同名會讓兩台攝影機的人流被靜默合併。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"], "cam002": ["entrance"]},
    )
    _write_zone_counts(output_dir / "zone_counts.parquet")

    with pytest.raises(ValueError, match="全域唯一"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_rejects_unknown_camera_zone_pair(tmp_path):
    """zone_counts.parquet 出現不在 camera_registry.yaml 定義內的
    (camera, zone) 組合時應 fail-loud，而非靜默讀入未經驗證的資料。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    # parquet 裡的 zone 是 "checkout"，但 registry 的 cam001 只定義了 "entrance"
    _write_zone_counts(output_dir / "zone_counts.parquet", zone="checkout")
    _write_registry(
        bucket_dir / "camera_registry.yaml", {"cam001": ["entrance"]}
    )

    with pytest.raises(ValueError, match="不在.*定義"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_skips_lines_when_registry_defines_none(tmp_path):
    """registry 沒有任何計數線定義時，缺 line_counts.parquet 不是錯誤——這個 bucket
    本來就沒有計數線，出入口三頁沒有資料可寫（分頁仍會建立）。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_zone_counts(output_dir / "zone_counts.parquet")
    _write_registry(
        bucket_dir / "camera_registry.yaml", {"cam001": ["entrance"]}
    )

    frames = _build(tmp_path, bucket_dir)
    assert frames.zone_hourly.height == 1
    assert frames.line_hourly is None
    assert frames.line_peak_in is None
    assert frames.line_peak_out is None


def test_build_report_frames_skips_zones_when_registry_defines_none(tmp_path):
    """反向亦然：registry 只定義計數線時，缺 zone_counts.parquet 不是錯誤。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_line_counts(output_dir / "line_counts.parquet")
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        lines_by_camera={"cam001": ["main_gate"]},
    )

    frames = _build(tmp_path, bucket_dir)
    assert frames.zone_hourly is None
    assert frames.zone_peak is None
    assert frames.line_hourly.height == 1


def test_build_report_frames_requires_line_counts_when_registry_defines_lines(tmp_path):
    """registry 定義了計數線卻沒有 line_counts.parquet：這是「忘了跑 line_counting」，
    不是「本來就沒有計數線」，必須 fail-loud 而非靜默少三頁。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_zone_counts(output_dir / "zone_counts.parquet")
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"]},
        {"cam001": ["main_gate"]},
    )

    with pytest.raises(FileNotFoundError, match="count_lines_daily"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_requires_zone_counts_when_registry_defines_zones(tmp_path):
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_line_counts(output_dir / "line_counts.parquet")
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"]},
        {"cam001": ["main_gate"]},
    )

    with pytest.raises(FileNotFoundError, match="map_zones_daily"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_rejects_registry_without_any_definition(tmp_path):
    """registry 裡既沒有區域也沒有計數線：沒有任何東西可彙總，不該產出空報表。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_registry(bucket_dir / "camera_registry.yaml", {"cam001": []})

    with pytest.raises(ValueError, match="沒有可彙總的統計"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_rejects_unknown_camera_line_pair(tmp_path):
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    # parquet 裡的計數線是 "side_gate"，但 registry 的 cam001 只定義了 "main_gate"
    _write_line_counts(output_dir / "line_counts.parquet", line="side_gate")
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        lines_by_camera={"cam001": ["main_gate"]},
    )

    with pytest.raises(ValueError, match="不在.*定義"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_rejects_duplicate_line_names_across_cameras(tmp_path):
    """計數線名稱跨攝影機重複要擋下（區域沒問題也一樣）：下游依計數線名稱、不含
    camera_id 分組彙總，同名會讓兩台攝影機的進出人數被靜默合併。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_zone_counts(output_dir / "zone_counts.parquet")
    _write_line_counts(output_dir / "line_counts.parquet")
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"]},
        {"cam001": ["main_gate"], "cam002": ["main_gate"]},
    )

    with pytest.raises(ValueError, match="計數線名稱.*全域唯一"):
        _build(tmp_path, bucket_dir)


def test_build_report_frames_accepts_empty_line_counts(tmp_path):
    """當日無任何跨越事件時 line_counts.parquet 是 0 列——那是 line_counting 的
    正常產物，不是缺資料，不可報錯。"""
    bucket_dir, output_dir = _prepare_bucket(tmp_path)
    _write_line_counts(output_dir / "line_counts.parquet", empty=True)
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        lines_by_camera={"cam001": ["main_gate"]},
    )

    frames = _build(tmp_path, bucket_dir)
    assert frames.line_hourly.height == 0
    assert frames.line_peak_in.height == 0
    assert frames.line_peak_out.height == 0


def _make_zone_hourly_df(rows):
    return pl.DataFrame(
        rows,
        schema={
            "date": pl.Utf8,
            "weekday": pl.Utf8,
            "period": pl.Utf8,
            "zone": pl.Utf8,
            "value": pl.Int64,
        },
        orient="row",
    )


def _make_zone_peak_df(rows):
    return pl.DataFrame(
        rows,
        schema={
            "date": pl.Utf8,
            "weekday": pl.Utf8,
            "zone": pl.Utf8,
            "peak_period": pl.Utf8,
            "peak_value": pl.Int64,
            "reminder": pl.Utf8,
        },
        orient="row",
    )


def _make_line_hourly_df(rows):
    return pl.DataFrame(
        rows,
        schema={
            "date": pl.Utf8,
            "weekday": pl.Utf8,
            "period": pl.Utf8,
            "line_group": pl.Utf8,
            "line": pl.Utf8,
            "in_count": pl.Int64,
            "out_count": pl.Int64,
            "net": pl.Int64,
        },
        orient="row",
    )


def _make_line_peak_df(rows):
    return pl.DataFrame(
        rows,
        schema={
            "date": pl.Utf8,
            "weekday": pl.Utf8,
            "line_group": pl.Utf8,
            "line": pl.Utf8,
            "peak_period": pl.Utf8,
            "peak_in": pl.Int64,
            "peak_out": pl.Int64,
            "reminder": pl.Utf8,
        },
        orient="row",
    )


def _frames(**kwargs) -> ReportFrames:
    fields = dict.fromkeys(ReportFrames._fields, None)
    fields.update(kwargs)
    return ReportFrames(**fields)


def _zone_frames(date="2026-05-01", weekday="星期五", period="09:00", value=10):
    return _frames(
        zone_hourly=_make_zone_hourly_df([(date, weekday, period, "checkout", value)]),
        zone_peak=_make_zone_peak_df(
            [(date, weekday, "checkout", period, value, "無")]
        ),
    )


def _full_frames(date="2026-05-01", weekday="星期五", period="09:00"):
    zone = _zone_frames(date=date, weekday=weekday, period=period)
    line_hourly = _make_line_hourly_df(
        [
            (date, weekday, period, "四樓書店", "main_gate", 30, 12, 18),
            (date, weekday, period, "四樓書店", "side_gate", 5, 9, -4),
        ]
    )
    peak_rows = [
        (date, weekday, "四樓書店", "main_gate", period, 30, 12, "無"),
        (date, weekday, "四樓書店", "side_gate", period, 5, 9, "無"),
    ]
    return zone._replace(
        line_hourly=line_hourly,
        line_peak_in=_make_line_peak_df(peak_rows),
        line_peak_out=_make_line_peak_df(peak_rows),
    )


def _write(path: Path, frames: ReportFrames, on_duplicate_date="append", date=None):
    """呼叫 _write_report 的測試捷徑：date 預設取 frames 帶到的日期。

    date 只在「兩側都沒有資料列」時才與 frames 的內容不同（見
    test_write_report_overwrite_clears_target_date_without_any_rows）。
    """
    if date is None:
        dates = {
            d
            for field in ReportFrames._fields
            if (df := getattr(frames, field)) is not None
            for d in df["date"].to_list()
        }
        assert len(dates) == 1, f"測試 frames 應只帶單一日期，實得 {sorted(dates)}"
        date = dates.pop()
    _write_report(path, frames, datetime.date.fromisoformat(date), on_duplicate_date)


def _sheet_rows(path: Path) -> dict[str, list[tuple]]:
    wb = openpyxl.load_workbook(path)
    rows = {
        ws.title: [tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        for ws in wb.worksheets
    }
    wb.close()
    return rows


def test_write_report_creates_all_sheets_with_headers(tmp_path):
    """不論該 bucket 有沒有計數線，6 個分頁的表頭一律建立，讓 BI 端的 schema 穩定。"""
    path = tmp_path / "report.xlsx"
    _write(path, _zone_frames(), on_duplicate_date="append")

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == list(_ALL_SHEETS)
    assert [c.value for c in wb[SHEET_ZONE_HOURLY][1]] == ZONE_HOURLY_HEADERS
    assert [c.value for c in wb[SHEET_ZONE_PEAK][1]] == ZONE_PEAK_HEADERS
    assert [c.value for c in wb[SHEET_LINE_HOURLY][1]] == LINE_HOURLY_HEADERS
    assert [c.value for c in wb[SHEET_LINE_PEAK_IN][1]] == LINE_PEAK_HEADERS
    assert [c.value for c in wb[SHEET_LINE_PEAK_OUT][1]] == LINE_PEAK_HEADERS
    assert [c.value for c in wb[SHEET_EVENTS][1]] == EVENTS_HEADERS
    # line 側為 None：分頁在、但沒有資料列
    assert wb[SHEET_LINE_HOURLY].max_row == 1
    wb.close()


def test_write_report_adds_missing_sheets_to_legacy_workbook(tmp_path):
    """既有的舊格式 report.xlsx：缺的分頁要補建，且不可生出「活動事件1」——
    openpyxl 的 create_sheet 對同名分頁不報錯，會靜默改名。"""
    path = tmp_path / "report.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    legacy_hourly = wb.create_sheet("每小時人流")
    legacy_hourly.append(ZONE_HOURLY_HEADERS)
    legacy_hourly.append(["2026-04-01", "星期三", "09:00", "checkout", 3])
    events = wb.create_sheet(SHEET_EVENTS)
    events.append(EVENTS_HEADERS)
    wb.save(path)
    wb.close()

    _write(path, _full_frames(), on_duplicate_date="append")

    wb = openpyxl.load_workbook(path)
    assert "活動事件1" not in wb.sheetnames
    assert wb.sheetnames.count(SHEET_EVENTS) == 1
    assert set(_ALL_SHEETS).issubset(wb.sheetnames)
    # 舊分頁留著不動、不再寫入
    legacy = [
        tuple(row) for row in wb["每小時人流"].iter_rows(min_row=2, values_only=True)
    ]
    assert legacy == [("2026-04-01", "星期三", "09:00", "checkout", 3)]
    wb.close()


def test_write_report_overwrite_removes_date_typed_existing_rows(tmp_path):
    """Excel／BI 工具開啟存檔後，日期欄的儲格可能被轉成 datetime.date 型別；
    overwrite 模式下 _existing_dates／_remove_rows_for_dates 仍須能辨識出目標
    日期並正確刪除舊列，不因型別不同（date vs str）而比對永遠不成立。"""
    path = tmp_path / "report.xlsx"
    _write(path, _zone_frames(period="09:00", value=10), "append")

    # 模擬用 Excel 開啟存檔後，日期欄的儲格被轉成 datetime.date 型別
    wb = openpyxl.load_workbook(path)
    for sheet_name in (SHEET_ZONE_HOURLY, SHEET_ZONE_PEAK):
        wb[sheet_name].cell(row=2, column=1).value = datetime.date(2026, 5, 1)
    wb.save(path)
    wb.close()

    _write(path, _zone_frames(period="10:00", value=20), "overwrite")

    hourly_rows = _sheet_rows(path)[SHEET_ZONE_HOURLY]
    # 舊列（09:00／10）已被覆蓋刪除，不是附加成第二列
    assert len(hourly_rows) == 1
    assert hourly_rows[0][4] == 20


def test_write_report_overwrite_sorts_mixed_date_types_without_crashing(tmp_path):
    """未被本次 overwrite 觸及的既有列可能仍是 datetime.date 型別（Excel 存檔
    造成），與本次新寫入的 str 型別日期混雜時，_sort_rows 排序不應因型別不同
    而 TypeError。"""
    path = tmp_path / "report.xlsx"
    _write(path, _zone_frames(date="2026-04-01", weekday="星期三"), "append")

    wb = openpyxl.load_workbook(path)
    for sheet_name in (SHEET_ZONE_HOURLY, SHEET_ZONE_PEAK):
        wb[sheet_name].cell(row=2, column=1).value = datetime.date(2026, 4, 1)
    wb.save(path)
    wb.close()

    # overwrite 目標是 2026-05-01，2026-04-01 不受影響、維持 date 型別
    _write(path, _zone_frames(period="10:00", value=20), "overwrite")

    dates = [
        d.strftime("%Y-%m-%d") if isinstance(d, datetime.date) else d
        for d, *_ in _sheet_rows(path)[SHEET_ZONE_HOURLY]
    ]
    assert dates == ["2026-04-01", "2026-05-01"]


def test_write_report_sorts_sheets_containing_blank_rows(tmp_path):
    """既有分頁可能含全空列（實測既有 report.xlsx 的活動事件分頁就有）；本次新增
    三個會被排序的分頁，_sort_key 拿 None 與 str 比較會 TypeError。"""
    path = tmp_path / "report.xlsx"
    _write(path, _full_frames(), "append")

    # 插在資料列之間才會被存檔保留（openpyxl 會裁掉檔尾的空列）
    wb = openpyxl.load_workbook(path)
    wb[SHEET_LINE_HOURLY].insert_rows(2)
    wb.save(path)
    wb.close()

    _write(path, _full_frames(date="2026-05-02", weekday="星期六"), "overwrite")

    rows = _sheet_rows(path)[SHEET_LINE_HOURLY]
    # 全空列排在最前面且被保留，兩天的資料都在
    assert rows[0][0] is None
    assert {row[0] for row in rows[1:]} == {"2026-05-01", "2026-05-02"}


def test_write_report_overwrite_is_idempotent_across_all_sheets(tmp_path):
    """同一天以 overwrite 重跑，6 個分頁的列序與內容都須逐列一致。"""
    path = tmp_path / "report.xlsx"
    _write(path, _full_frames(), "overwrite")
    first = _sheet_rows(path)
    _write(path, _full_frames(), "overwrite")

    assert _sheet_rows(path) == first


def test_write_report_overwrite_clears_line_sheets_when_line_side_is_none(tmp_path):
    """registry 移除 lines 後以 overwrite 重跑該日：出入口三頁的舊列要一併清除，
    否則同一天會在不同分頁混雜新舊資料。"""
    path = tmp_path / "report.xlsx"
    _write(path, _full_frames(), "append")
    assert _sheet_rows(path)[SHEET_LINE_HOURLY]

    _write(path, _zone_frames(period="10:00", value=99), "overwrite")

    rows = _sheet_rows(path)
    assert rows[SHEET_LINE_HOURLY] == []
    assert rows[SHEET_LINE_PEAK_IN] == []
    assert rows[SHEET_LINE_PEAK_OUT] == []
    assert len(rows[SHEET_ZONE_HOURLY]) == 1


def test_write_report_overwrite_clears_target_date_without_any_rows(tmp_path):
    """上游重跑後該日事件清空（0 列，不是缺資料）時 overwrite 仍須清掉該日舊列：
    要清的日期取自本次彙總的 date，只看資料內容的話 0 列就清不到任何東西。"""
    path = tmp_path / "report.xlsx"
    _write(path, _full_frames(), "append")
    assert _sheet_rows(path)[SHEET_ZONE_HOURLY]

    empty = _frames(
        zone_hourly=_make_zone_hourly_df([]),
        zone_peak=_make_zone_peak_df([]),
        line_hourly=_make_line_hourly_df([]),
        line_peak_in=_make_line_peak_df([]),
        line_peak_out=_make_line_peak_df([]),
    )
    _write(path, empty, "overwrite", date="2026-05-01")

    rows = _sheet_rows(path)
    assert all(rows[sheet] == [] for sheet in _ALL_SHEETS)


def test_write_report_error_mode_writes_nothing_on_conflict(tmp_path):
    """error 模式遇到日期衝突時，6 個分頁都不可被寫入（跨分頁的原子性）。"""
    path = tmp_path / "report.xlsx"
    _write(path, _full_frames(), "append")
    before = _sheet_rows(path)

    with pytest.raises(ValueError, match="未寫入任何內容"):
        _write(path, _full_frames(period="10:00"), "error")

    assert _sheet_rows(path) == before


def test_write_report_error_mode_detects_conflict_on_line_sheets(tmp_path):
    """衝突偵測涵蓋 line 分頁：zone 那側沒有衝突也要擋下。"""
    path = tmp_path / "report.xlsx"
    line_only = _frames(
        line_hourly=_make_line_hourly_df(
            [("2026-05-01", "星期五", "09:00", "四樓書店", "main_gate", 3, 1, 2)]
        )
    )
    _write(path, line_only, "append")

    with pytest.raises(ValueError, match="未寫入任何內容"):
        _write(path, line_only, "error")


def test_write_report_closes_workbook_after_save(tmp_path):
    """_write_report 開啟的 Workbook 用畢須 close，避免底層檔案控制代碼不釋放。"""
    path = tmp_path / "report.xlsx"

    with mock.patch(
        "openpyxl.workbook.workbook.Workbook.close", autospec=True
    ) as mock_close:
        _write(path, _zone_frames(), on_duplicate_date="append")

    assert mock_close.called
