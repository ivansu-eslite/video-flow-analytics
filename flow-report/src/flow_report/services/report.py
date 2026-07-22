"""Zone 人流 Excel 報表：核心匯出邏輯（讀檔、驗證、orchestration 與 Excel 讀寫）。

讀 `outputs/{bucket}/{date}/zone_counts.parquet`，彙總成跨日累加更新的
`outputs/{bucket}/report.xlsx`。實際的期間彙總／尖峰計算在 `services/stats.py`。
"""

import datetime
from pathlib import Path
from typing import Literal

import openpyxl
import polars as pl
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from flow_report.config.constants import (
    COLUMN_WIDTH,
    EVENTS_HEADERS,
    HOURLY_HEADERS,
    HOURLY_SORT_COLUMNS,
    OUTPUT_ROOT,
    PEAK_HEADERS,
    PEAK_SORT_COLUMNS,
    REGISTRY_SNAPSHOT_FILENAME,
    REPORT_FILENAME,
    SHEET_EVENTS,
    SHEET_HOURLY,
    SHEET_PEAK,
    TMP_SUFFIX,
    ZONE_COUNTS_FILENAME,
)
from flow_report.models.registry import (
    load_registry_from_path,
    parse_and_validate_zones,
)
from flow_report.observability import StructuredLogger
from flow_report.services.stats import peak_per_day, rollup_by_period, to_taipei

logger = StructuredLogger(component="report_builder")


def _build_report_frames(
    date: datetime.date,
    bucket_dir: str,
    period_minutes: int,
    metric: str,
    bucket_minutes: int,
    output_root: Path = OUTPUT_ROOT,
) -> tuple[pl.DataFrame, pl.DataFrame]:
    if period_minutes % bucket_minutes != 0:
        raise ValueError(
            f"report.period_minutes（{period_minutes}）必須是 "
            f"zone.bucket_minutes（{bucket_minutes}）的倍數。"
        )

    bucket_path = Path(bucket_dir)
    output_dir = output_root / bucket_path.name / date.isoformat()
    counts_path = output_dir / ZONE_COUNTS_FILENAME
    if not counts_path.exists():
        raise FileNotFoundError(
            f"找不到 zone 人流統計 {counts_path}，"
            "請先執行 map_zones_daily 產生當日 parquet。"
        )

    # 為何用快照而非當下 registry：見 export_report_daily 的 docstring 說明
    registry = load_registry_from_path(output_dir / REGISTRY_SNAPSHOT_FILENAME)
    zone_entries = {
        entry.stream_dirname: entry
        for entry in registry.cameras
        if entry.participates_in_zone_mapping
    }
    # parse_and_validate_zones 順便驗證跨攝影機 zone 名稱唯一性
    zone_cameras = parse_and_validate_zones(zone_entries)

    df = to_taipei(pl.read_parquet(counts_path))
    valid_pairs = {
        (camera_id, zone.name)
        for camera_id, zones in zone_cameras.items()
        for zone in zones
    }
    actual_pairs = set(zip(df["camera_id"].to_list(), df["zone"].to_list()))
    unknown_pairs = actual_pairs - valid_pairs
    if unknown_pairs:
        raise ValueError(
            f"{counts_path} 出現不在 camera_registry_used.yaml 快照內的 "
            f"(camera, zone) 組合: {sorted(unknown_pairs)}"
        )

    hourly_df = rollup_by_period(df, period_minutes, metric)
    peak_df = peak_per_day(hourly_df)
    return hourly_df, peak_df


def _sort_key_columns(headers: list[str], columns: tuple[str, ...]) -> tuple[int, ...]:
    return tuple(headers.index(column) for column in columns)


def _init_sheet(wb: Workbook, name: str, headers: list[str]) -> Worksheet:
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = COLUMN_WIDTH
    return ws


def _cell_date_str(value: object) -> str | None:
    """把日期欄的儲格值正規化成 `YYYY-MM-DD` 字串。

    本階段以字串寫入日期，但 Excel／BI 工具存檔時可能把該欄轉成日期型別的儲格，
    讀回時即為 `datetime.date`／`datetime.datetime`。日期欄同時是比對與排序的鍵，
    型別混雜會讓比對永遠不成立、排序直接拋 `TypeError`，故一律正規化後再使用。
    """
    if value is None:
        return None
    if isinstance(value, datetime.date):  # datetime.datetime 亦為其子類
        return value.strftime("%Y-%m-%d")
    return str(value)


def _existing_dates(ws: Worksheet) -> set[str]:
    dates = (_cell_date_str(row[0].value) for row in ws.iter_rows(min_row=2))
    return {date for date in dates if date is not None}


def _remove_rows_for_dates(ws: Worksheet, dates: set[str]) -> None:
    rows_to_delete = [
        row[0].row
        for row in ws.iter_rows(min_row=2)
        if _cell_date_str(row[0].value) in dates
    ]
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)


def _append_rows(ws: Worksheet, df: pl.DataFrame) -> None:
    for row in df.iter_rows():
        ws.append(row)


def _sort_key(value: object) -> object:
    """排序鍵正規化：日期型別的儲格轉字串，其餘型別原樣保留。

    key_columns 可能包含非日期欄（如區域名稱），只正規化日期型別可避免混入
    `datetime.date`／`str` 時排序互相比較拋 `TypeError`，同時不影響其他欄位
    的原生型別比較。
    """
    if isinstance(value, datetime.date):  # datetime.datetime 亦為其子類
        return _cell_date_str(value)
    return value


def _sort_rows(ws: Worksheet, key_columns: tuple[int, ...]) -> None:
    if ws.max_row < 2:
        return
    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    rows.sort(key=lambda r: tuple(_sort_key(r[i]) for i in key_columns))
    ws.delete_rows(2, ws.max_row - 1)
    for row in rows:
        ws.append(row)


def _write_report(
    path: Path,
    hourly_new: pl.DataFrame,
    peak_new: pl.DataFrame,
    on_duplicate_date: Literal["overwrite", "append", "error"],
) -> None:
    new_dates = set(hourly_new["date"].to_list())

    if path.exists():
        wb = openpyxl.load_workbook(path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        _init_sheet(wb, SHEET_HOURLY, HOURLY_HEADERS)
        _init_sheet(wb, SHEET_PEAK, PEAK_HEADERS)
        _init_sheet(wb, SHEET_EVENTS, EVENTS_HEADERS)
        wb.remove(default_sheet)

    try:
        hourly_ws = wb[SHEET_HOURLY]
        peak_ws = wb[SHEET_PEAK]

        if on_duplicate_date == "error":
            conflict = new_dates & (
                _existing_dates(hourly_ws) | _existing_dates(peak_ws)
            )
            if conflict:
                raise ValueError(
                    f"報表中已存在這些日期的資料，未寫入任何內容：{sorted(conflict)}"
                    "（可改用 on_duplicate_date='overwrite' 或 'append'）"
                )

        if on_duplicate_date == "overwrite":
            _remove_rows_for_dates(hourly_ws, new_dates)
            _remove_rows_for_dates(peak_ws, new_dates)

        _append_rows(hourly_ws, hourly_new)
        _append_rows(peak_ws, peak_new)

        if on_duplicate_date == "overwrite":
            _sort_rows(
                hourly_ws,
                key_columns=_sort_key_columns(HOURLY_HEADERS, HOURLY_SORT_COLUMNS),
            )
            _sort_rows(
                peak_ws,
                key_columns=_sort_key_columns(PEAK_HEADERS, PEAK_SORT_COLUMNS),
            )

        tmp_path = path.with_name(path.name + TMP_SUFFIX)
        wb.save(tmp_path)
        tmp_path.replace(path)
    finally:
        wb.close()


def export_report_daily(
    date: datetime.date,
    bucket_dir: str,
    period_minutes: int,
    metric: Literal["entries", "unique_visitors"],
    on_duplicate_date: Literal["overwrite", "append", "error"],
    bucket_minutes: int,
    output_root: Path = OUTPUT_ROOT,
) -> Path:
    """執行單日 zone 人流報表彙總，寫入跨日累加更新的 `report.xlsx`。

    純 CPU 運算，不需重跑偵測或 zone mapping；讀取的 registry 是產生
    `zone_counts.parquet` 當時的 `camera_registry_used.yaml` 快照（而非
    當下的 `camera_registry.yaml`），以避免同名 zone 被靜默合併。

    Args:
        date: 要彙總的日期，需已有對應的 `zone_counts.parquet`。
        bucket_dir: 本機模擬 GCS bucket 的根目錄。
        period_minutes: 報表人流彙總的時段粒度（分鐘），需為 `bucket_minutes`
            的倍數。
        metric: 「人流量」「尖峰人流」使用的統計量。
        on_duplicate_date: 同一天資料已存在時的處理方式。
        bucket_minutes: `zone_counts.parquet` 的時段粒度（分鐘）。
        output_root: 輸出根目錄。

    Returns:
        `report.xlsx` 的路徑。

    Raises:
        ValueError: `period_minutes` 不是 `bucket_minutes` 的倍數、
            `camera_registry_used.yaml` 中有跨攝影機重複的 zone 名稱、
            `zone_counts.parquet` 出現不在該快照內的 (camera, zone) 組合，或
            `on_duplicate_date="error"` 時發現日期已存在。
        FileNotFoundError: 當日 `zone_counts.parquet` 不存在，或該日輸出
            目錄下找不到 `camera_registry_used.yaml` 快照。
    """
    hourly_df, peak_df = _build_report_frames(
        date, bucket_dir, period_minutes, metric, bucket_minutes, output_root
    )

    bucket_name = Path(bucket_dir).name
    report_path = output_root / bucket_name / REPORT_FILENAME
    _write_report(report_path, hourly_df, peak_df, on_duplicate_date)

    logger.info(
        "Zone 人流報表已寫入",
        path=str(report_path),
        dates=sorted(hourly_df["date"].unique().to_list()),
        hourly_rows=hourly_df.height,
        peak_rows=peak_df.height,
    )
    return report_path
