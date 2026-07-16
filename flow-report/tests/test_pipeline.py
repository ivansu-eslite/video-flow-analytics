import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import polars as pl
import pytest
import yaml

from flow_report.pipeline import SHEET_HOURLY, SHEET_PEAK, _build_report_frames, _write_report


def _write_registry(path: Path, zones_by_camera: dict[str, list[str]]) -> None:
    data = {
        "bucket_name": "bucket_test",
        "storage": {},
        "cameras": [
            {
                "camera_id": cam_id,
                "location": "loc",
                "ip": "127.0.0.1",
                "zones": [
                    {"name": name, "polygon": [[0, 0], [1, 0], [1, 1]]}
                    for name in names
                ],
            }
            for cam_id, names in zones_by_camera.items()
        ],
    }
    path.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False), encoding="utf-8"
    )


def _write_zone_counts(path: Path) -> None:
    df = pl.DataFrame(
        {
            "camera_id": ["cam001"],
            "zone": ["entrance"],
            "time_bucket": [
                datetime.datetime(2026, 5, 1, 11, 0, tzinfo=ZoneInfo("Asia/Taipei"))
            ],
            "unique_visitors": [1],
            "entries": [1],
        },
        schema={
            "camera_id": pl.Utf8,
            "zone": pl.Utf8,
            "time_bucket": pl.Datetime("us", "Asia/Taipei"),
            "unique_visitors": pl.Int64,
            "entries": pl.Int64,
        },
    )
    df.write_parquet(path)


def test_build_report_frames_validates_snapshot_registry_not_live(tmp_path):
    """_build_report_frames 應該驗證產生 zone_counts.parquet 當時的 registry 快照
    （camera_registry_used.yaml），而不是「當下」的 camera_registry.yaml。"""
    bucket_dir = tmp_path / "bucket_test"
    bucket_dir.mkdir()
    # 即時 registry：已經修正成不重複
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"], "cam002": ["entrance_2"]},
    )

    output_dir = tmp_path / "outputs" / "bucket_test" / "2026-05-01"
    output_dir.mkdir(parents=True)
    _write_zone_counts(output_dir / "zone_counts.parquet")
    # 快照：產生 parquet 當時兩台攝影機都叫 entrance（重複）
    _write_registry(
        output_dir / "camera_registry_used.yaml",
        {"cam001": ["entrance"], "cam002": ["entrance"]},
    )

    with pytest.raises(ValueError, match="全域唯一"):
        _build_report_frames(
            date=datetime.date(2026, 5, 1),
            bucket_dir=str(bucket_dir),
            period_minutes=60,
            metric="entries",
            bucket_minutes=15,
            output_root=tmp_path / "outputs",
        )


def test_build_report_frames_ignores_live_registry_duplicates(tmp_path):
    """即時 camera_registry.yaml 目前有重複也不該擋下報表，只要產生資料當時的
    快照是唯一的就該正常執行。"""
    bucket_dir = tmp_path / "bucket_test"
    bucket_dir.mkdir()
    # 即時 registry 目前有重複（尚未修正），但不該影響驗證結果
    _write_registry(
        bucket_dir / "camera_registry.yaml",
        {"cam001": ["entrance"], "cam002": ["entrance"]},
    )

    output_dir = tmp_path / "outputs" / "bucket_test" / "2026-05-01"
    output_dir.mkdir(parents=True)
    _write_zone_counts(output_dir / "zone_counts.parquet")
    # 快照：產生 parquet 當時是全域唯一的
    _write_registry(
        output_dir / "camera_registry_used.yaml",
        {"cam001": ["entrance"], "cam002": ["entrance_2"]},
    )

    hourly_df, peak_df = _build_report_frames(
        date=datetime.date(2026, 5, 1),
        bucket_dir=str(bucket_dir),
        period_minutes=60,
        metric="entries",
        bucket_minutes=15,
        output_root=tmp_path / "outputs",
    )
    assert hourly_df.height == 1
    assert peak_df.height == 1


def _make_hourly_df(rows):
    return pl.DataFrame(
        rows,
        schema={
            "date": pl.Utf8,
            "weekday": pl.Utf8,
            "period": pl.Utf8,
            "zone": pl.Utf8,
            "value": pl.Int64,
        },
        orient="row",
    )


def _make_peak_df(rows):
    return pl.DataFrame(
        rows,
        schema={
            "date": pl.Utf8,
            "weekday": pl.Utf8,
            "zone": pl.Utf8,
            "peak_period": pl.Utf8,
            "peak_value": pl.Int64,
            "reminder": pl.Utf8,
        },
        orient="row",
    )


def test_write_report_overwrite_removes_date_typed_existing_rows(tmp_path):
    """Excel／BI 工具開啟存檔後，日期欄的儲格可能被轉成 datetime.date 型別；
    overwrite 模式下 _existing_dates／_remove_rows_for_dates 仍須能辨識出目標
    日期並正確刪除舊列，不因型別不同（date vs str）而比對永遠不成立。"""
    path = tmp_path / "report.xlsx"
    hourly_1 = _make_hourly_df([("2026-05-01", "星期五", "09:00", "checkout", 10)])
    peak_1 = _make_peak_df([("2026-05-01", "星期五", "checkout", "09:00", 10, "無")])
    _write_report(path, hourly_1, peak_1, on_duplicate_date="append")

    # 模擬用 Excel 開啟存檔後，日期欄的儲格被轉成 datetime.date 型別
    wb = openpyxl.load_workbook(path)
    for sheet_name in (SHEET_HOURLY, SHEET_PEAK):
        wb[sheet_name].cell(row=2, column=1).value = datetime.date(2026, 5, 1)
    wb.save(path)
    wb.close()

    hourly_2 = _make_hourly_df([("2026-05-01", "星期五", "10:00", "checkout", 20)])
    peak_2 = _make_peak_df([("2026-05-01", "星期五", "checkout", "10:00", 20, "無")])
    _write_report(path, hourly_2, peak_2, on_duplicate_date="overwrite")

    result = openpyxl.load_workbook(path)
    hourly_rows = [
        tuple(row)
        for row in result[SHEET_HOURLY].iter_rows(min_row=2, values_only=True)
    ]
    # 舊列（09:00／10）已被覆蓋刪除，不是附加成第二列
    assert len(hourly_rows) == 1
    assert hourly_rows[0][4] == 20
    result.close()


def test_write_report_overwrite_sorts_mixed_date_types_without_crashing(tmp_path):
    """未被本次 overwrite 觸及的既有列可能仍是 datetime.date 型別（Excel 存檔
    造成），與本次新寫入的 str 型別日期混雜時，_sort_rows 排序不應因型別不同
    而 TypeError。"""
    path = tmp_path / "report.xlsx"
    hourly_1 = _make_hourly_df([("2026-04-01", "星期三", "09:00", "checkout", 3)])
    peak_1 = _make_peak_df([("2026-04-01", "星期三", "checkout", "09:00", 3, "無")])
    _write_report(path, hourly_1, peak_1, on_duplicate_date="append")

    wb = openpyxl.load_workbook(path)
    for sheet_name in (SHEET_HOURLY, SHEET_PEAK):
        wb[sheet_name].cell(row=2, column=1).value = datetime.date(2026, 4, 1)
    wb.save(path)
    wb.close()

    # overwrite 目標是 2026-05-01，2026-04-01 不受影響、維持 date 型別
    hourly_2 = _make_hourly_df([("2026-05-01", "星期五", "10:00", "checkout", 20)])
    peak_2 = _make_peak_df([("2026-05-01", "星期五", "checkout", "10:00", 20, "無")])
    _write_report(path, hourly_2, peak_2, on_duplicate_date="overwrite")

    result = openpyxl.load_workbook(path)
    hourly_rows = [
        tuple(row)
        for row in result[SHEET_HOURLY].iter_rows(min_row=2, values_only=True)
    ]
    dates = [
        d.strftime("%Y-%m-%d") if isinstance(d, datetime.date) else d
        for d, *_ in hourly_rows
    ]
    assert dates == ["2026-04-01", "2026-05-01"]
    result.close()
